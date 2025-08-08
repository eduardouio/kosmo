from io import BytesIO
from datetime import date
import unicodedata
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
import os
from django.conf import settings

from .PartnerAccountStatmentView import PartnerAccountStatmentView


class PartnerAccountStatmentExcel(View):
    """Exporta el estado de cuenta a Excel (xlsx)."""

    def get(self, request, *args, **kwargs):
        base_view = PartnerAccountStatmentView()
        base_view.request = request
        ctx = base_view.get_context_data()

        if ctx.get('missing_partner') or ctx.get('partner_not_found'):
            return HttpResponse(
                'Debe seleccionar un cliente válido.', status=400
            )

        wb = Workbook()
        ws = wb.active
        ws.title = 'EstadoCuenta'

        partner = ctx['partner']
        # Logo (si existe)
        logo_paths = [
            os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-kosmo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-bg.png'),
        ]
        placed_logo = False
        for lp in logo_paths:
            if os.path.exists(lp):
                try:
                    img = XLImage(lp)
                    img.height = 55
                    img.width = 180
                    ws.add_image(img, 'A1')
                    placed_logo = True
                    break
                except Exception:
                    continue

        title_row = 1 if placed_logo else 1
        ws.append(['ESTADO DE CUENTA'])
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=7)
        ws['A{}'.format(title_row)].font = Font(size=14, bold=True)
        # Aumentar altura para que el logo no se sobreponga con el título
        ws.row_dimensions[title_row].height = 65

        ws.append([
            f'Cliente: {partner.name}', '', '', '', 'Fecha',
            date.today().strftime('%d/%m/%Y')
        ])
        ws.append([
            f'Tax ID: {partner.business_tax_id or "-"}', '', '', '', 'Saldo',
            ctx['net_balance']
        ])
        ws.append([f'Crédito: {partner.credit_term} días'])
        ws.append([])

        headers = [
            'Fecha', 'Documento', 'Crédito', 'Valor', 'Pago', 'Saldo',
            'Referencia'
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(
                start_color='F6DFD4', end_color='F6DFD4', fill_type='solid'
            )

        row = 6
        for e in ctx['entries']:
            fecha = e.get('date')
            ws.cell(
                row=row, column=1,
                value=fecha.strftime('%d/%m/%Y') if fecha else ''
            )
            ws.cell(row=row, column=2, value=e.get('document'))
            ws.cell(
                row=row, column=3,
                value=e.get('credit_days') if e['type'] == 'INVOICE' else ''
            )
            ws.cell(
                row=row, column=4,
                value=(
                    e.get('invoice_amount') if e['type'] == 'INVOICE' else None
                )
            )
            ws.cell(
                row=row, column=5,
                value=(
                    e.get('payment_amount') if e['type'] == 'PAYMENT'
                    else (
                        e.get('payment_amount')
                        if e.get('payment_amount') else None
                    )
                )
            )
            ws.cell(
                row=row, column=6,
                value=(
                    e.get('balance') if e['type'] == 'INVOICE' else None
                )
            )
            ws.cell(row=row, column=7, value=e.get('reference'))
            row += 1

        ws.append([])
        ws.append([
            'TOTAL FACTURAS', ctx['total_invoices_amount'],
            'TOTAL PAGOS', ctx['total_payments_amount'],
            'PENDIENTE', ctx['total_pending_balance']
        ])

        # Nombre de archivo dinámico con el nombre del partner
        def slugify(value: str) -> str:
            value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
            value = value.lower().replace(' ', '_')
            return ''.join(c for c in value if c.isalnum() or c in ('_', '-')) or 'partner'

        filename = f"estado_cuenta_{slugify(partner.name)}.xlsx"
        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            )
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        response.write(bio.read())
        return response
